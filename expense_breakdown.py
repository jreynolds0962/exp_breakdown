# 1. figure out whether this is for all stores or just home outlet stores
# 2. figure out total amount of invoice
# 3. home outlet stores should be divided by number of stores - 1(becuase of home office account)
# 4. open up file
# 5. go to home outlet or all tab
# 6. fill in breakdown amount to amount column

from openpyxl import load_workbook




# 1.
def get_stores():
  stores = str(input("What stores? type 'ho', 'home outlet', or 'all': ").lower())
  if stores == 'ho' or stores == 'home outlet':
    return "Home Outlet"
  elif stores == "all":
    return "All"
  else:
    print("Error: Please type 'ho' 'home outlet', or 'all':")
    get_stores()

#1.5
def get_invoice_number():
  inv_num = input("What is the invoice reference? (include vendor name): ")
  return inv_num

# 2.
def get_total():
  try:
    total = float(input("What is the total of the invoice?: "))
    if type(total) is int or type(total) is float:
      answer = str(input("Are you sure $" + str(total) + " is the correct total?: ").lower())
      if answer == 'y' or answer == 'yes':
        return total
      elif answer == 'n' or answer == 'no':
        get_total()
      else:
        print("Your answer did not process correctly. Please input the total again")
        get_total()
  except:
    print("Type only the total amount")
    get_total()


inv_ref = get_invoice_number()
what_stores = get_stores()
invoice_total = get_total()

# 3.
if what_stores == "Home Outlet":
  num_of_stores = int(97)
elif what_stores == "All":
  num_of_stores = int(108)
else:
  print("Error, something went wrong.")

expense_distribution = round(invoice_total / num_of_stores, 2)
print(expense_distribution, "will go to each store")


# 4.
wb = load_workbook('import_spreadsheet.xlsx')


#5 & 6
if what_stores == 'Home Outlet':
  ws = wb['HOME OUTLET']
  for cell in range(1, 98):
    ws["C" + str(cell)] = expense_distribution
  for cell in range(1, 99):
    ws["A" + str(cell)] = inv_ref
elif what_stores == 'All':
  ws = wb['ALL STORES']
  for cell in range(1, 109):
    ws["C" + str(cell)] = expense_distribution
  for cell in range(1, 110):
    ws["A" + str(cell)] = inv_ref
else:
  ws = wb.active
  print("Did you choose a store range?")


# > and < operand types do not work on floats, must manually adjust values by .01 to reach invoice_total if there is a remainder

wb.save('import spreadsheet.xlsx')


